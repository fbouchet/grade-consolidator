"""
Test suite for consolidate_grades.
"""

import csv

import pandas as pd
import pytest
import yaml

from consolidate_grades.consolidate import (
    build_master_index,
    clean_column_names,
    consolidate,
    detect_all_columns,
    detect_column,
    detect_grade_column,
    find_header_row,
    levenshtein_distance,
    load_config,
    name_similarity_hint,
    normalize_text,
    parse_grade,
    process_ta_file,
    promote_header_row,
    prompt_column_choice,
    prompt_name_mismatch,
    prompt_sheet_selection,
    read_file,
    read_file_sheets,
    resolve_grade_files,
    write_moodle_csv,
)

# ============================================================================
# Helpers
# ============================================================================


def _write_csv(path, header, rows, sep=",", encoding="utf-8"):
    with open(path, "w", encoding=encoding, newline="") as f:
        w = csv.writer(f, delimiter=sep)
        w.writerow(header)
        w.writerows(rows)


def _master_df(students):
    """students = list of (id, first, last, email)."""
    return pd.DataFrame(
        {
            "Numéro étudiant": [s[0] for s in students],
            "Prénom": [s[1] for s in students],
            "Nom de famille": [s[2] for s in students],
            "Email": [s[3] for s in students],
        }
    )


def _build_master():
    """Standard 4-student master roster."""
    df = _master_df(
        [
            ("12345", "Jean", "Dupont", "jean@etu.fr"),
            ("12346", "Marie", "Curie", "marie@etu.fr"),
            ("12347", "Pierre", "Martin", "pierre@etu.fr"),
            ("12348", "Sophie", "Bernard", "sophie@etu.fr"),
        ]
    )
    master, _ = build_master_index(df)
    return master


# ============================================================================
# 1. normalize_text
# ============================================================================


class TestNormalizeText:
    def test_basic_lowercase(self):
        assert normalize_text("Hello") == "hello"

    def test_strip_whitespace(self):
        assert normalize_text("  hello  ") == "hello"

    def test_collapse_whitespace(self):
        assert normalize_text("hello   world") == "hello world"

    def test_accents_removed(self):
        assert normalize_text("Numéro étudiant") == "numero etudiant"

    def test_apostrophe_to_space(self):
        assert normalize_text("d'étudiant") == "d etudiant"

    def test_curly_apostrophe(self):
        assert normalize_text("d\u2019etudiant") == "d etudiant"

    def test_dashes_to_space(self):
        assert normalize_text("nom-de-famille") == "nom de famille"

    def test_en_dash(self):
        assert normalize_text("nom\u2013de\u2013famille") == "nom de famille"

    def test_em_dash(self):
        assert normalize_text("nom\u2014de\u2014famille") == "nom de famille"

    def test_underscore(self):
        assert normalize_text("nom_de_famille") == "nom de famille"

    def test_n_degree(self):
        assert normalize_text("N° étudiant") == "no etudiant"

    def test_n_ordinal_o(self):
        assert normalize_text("Nº étudiant") == "no etudiant"

    def test_c1_control_stripped(self):
        assert normalize_text("Num\u0092ro d\u0092identification") == "num ro d identification"

    def test_acute_accent_apostrophe(self):
        assert normalize_text("d\u00b4etudiant") == "d etudiant"

    def test_prime_apostrophe(self):
        assert normalize_text("d\u2032etudiant") == "d etudiant"

    def test_fullwidth_apostrophe(self):
        assert normalize_text("d\uff07etudiant") == "d etudiant"

    def test_n_degree_no_space(self):
        assert normalize_text("N°étudiant") == "no etudiant"
        assert normalize_text("N°identification") == "no identification"


class TestFullNameColumn:
    def test_nom_etu_detected(self):
        assert detect_column(["Nom étu", "Note"], "full_name") == "Nom étu"

    def test_full_pipeline_with_merged_name_and_id(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Nom étu", "N°étudiant", "TOTAL"],
            [["Jean Dupont", "12345", "15"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0

    def test_full_name_fallback_no_id(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Nom étu", "Note"],
            [["Jean Dupont", "15"], ["Curie Marie", "14"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.grades_assigned == 2
        assert master.by_id["12345"].grade == 15.0
        assert master.by_id["12346"].grade == 14.0


# ============================================================================
# 2. clean_column_names
# ============================================================================


class TestCleanColumnNames:
    def test_strip_bom(self):
        df = pd.DataFrame({"\ufeffNumero": [1]})
        df = clean_column_names(df)
        assert "Numero" in df.columns

    def test_strip_zero_width(self):
        df = pd.DataFrame({"\u200bPrenom": [1]})
        df = clean_column_names(df)
        assert "Prenom" in df.columns

    def test_strip_whitespace(self):
        df = pd.DataFrame({"  Note  ": [1]})
        df = clean_column_names(df)
        assert "Note" in df.columns


# ============================================================================
# 3. detect_column
# ============================================================================


class TestDetectColumn:
    @pytest.mark.parametrize(
        "col_name",
        [
            "Numéro étudiant",
            "numero etudiant",
            "NUMERO ETUDIANT",
            "Student ID",
            "student_id",
            "ID",
            "NIP",
            "Identifiant",
            "Code étudiant",
            "No étudiant",
            "Numéro d'identification",
            "N° d'identification",
            "Numero d'etudiant",
            "Numéro d'étudiant",
            "N° d'étudiant",
        ],
    )
    def test_id_variants(self, col_name):
        assert detect_column([col_name, "Other"], "id") == col_name

    @pytest.mark.parametrize(
        "col_name",
        ["Prénom", "prenom", "First Name", "firstname", "Given Name"],
    )
    def test_first_name_variants(self, col_name):
        assert detect_column([col_name, "Other"], "first_name") == col_name

    @pytest.mark.parametrize(
        "col_name",
        ["Nom", "Nom de famille", "Last Name", "Surname", "Family Name"],
    )
    def test_last_name_variants(self, col_name):
        assert detect_column([col_name, "Other"], "last_name") == col_name

    @pytest.mark.parametrize(
        "col_name",
        [
            "Note",
            "Notes",
            "Grade",
            "Score",
            "Résultat",
            "Mark",
            "Points",
            "Note finale",
            "Total",
        ],
    )
    def test_grade_variants(self, col_name):
        assert detect_column([col_name, "Other"], "grade") == col_name

    def test_no_match_returns_none(self):
        assert detect_column(["Foo", "Bar"], "id") is None

    def test_first_match_wins(self):
        assert detect_column(["ID", "Student ID"], "id") == "ID"


class TestDetectAllColumns:
    def test_single_match(self):
        assert detect_all_columns(["Numéro étudiant", "Note"], "id") == ["Numéro étudiant"]

    def test_multiple_matches(self):
        cols = ["Numero", "Prenom", "Numero etudiant", "Note"]
        matches = detect_all_columns(cols, "id")
        assert "Numero" in matches
        assert "Numero etudiant" in matches
        assert len(matches) == 2

    def test_no_match(self):
        assert detect_all_columns(["Foo", "Bar"], "id") == []


# ============================================================================
# 4. detect_grade_column
# ============================================================================


class TestDetectGradeColumn:
    def test_simple_note(self):
        df = pd.DataFrame({"ID": ["1"], "Note": ["15"]})
        col, _, _ = detect_grade_column(df, {"ID"})
        assert col == "Note"

    def test_prefix_match_note_slash_20(self):
        df = pd.DataFrame({"ID": ["1"], "Note /20": ["15"]})
        col, _warnings, _ = detect_grade_column(df, {"ID"})
        assert col == "Note /20"
        assert any("prefix" in w.lower() for w in _warnings)

    def test_prefix_match_note_slash_23(self):
        df = pd.DataFrame({"ID": ["1"], "Note /23": ["18"]})
        col, _, _ = detect_grade_column(df, {"ID"})
        assert col == "Note /23"

    def test_prefix_match_notes_sur(self):
        df = pd.DataFrame({"Numéro étudiant": ["1"], "Notes sur 23": ["15"]})
        col, _warnings, _ = detect_grade_column(df, {"Numéro étudiant"})
        assert col == "Notes sur 23"

    def test_total_preferred_over_questions(self):
        df = pd.DataFrame(
            {
                "ID": ["1"],
                "Q1": ["1"],
                "Q2": ["2"],
                "Total": ["15"],
            }
        )
        col, _, _ = detect_grade_column(df, {"ID"})
        assert col == "Total"

    def test_ambiguous_grades(self):
        df = pd.DataFrame({"ID": ["1"], "Note /20": ["15"], "Note /23": ["18"]})
        col, _, ambig = detect_grade_column(df, {"ID"})
        assert col is None
        assert "Note /20" in ambig
        assert "Note /23" in ambig

    def test_no_grade_column(self):
        df = pd.DataFrame({"ID": ["1"], "Foo": ["bar"]})
        col, _warnings, _ = detect_grade_column(df, {"ID"})
        assert col is None
        assert any("no grade" in w.lower() for w in _warnings)


# ============================================================================
# 5. find_header_row / promote_header_row
# ============================================================================


class TestFindHeaderRow:
    def test_normal_headers_no_promotion(self):
        df = pd.DataFrame({"Numéro étudiant": ["1"], "Prénom": ["Jean"]})
        assert find_header_row(df) is None

    def test_title_row_above_headers(self):
        # First column is "Title", others empty — current headers don't look right
        df = pd.DataFrame(
            {
                "Title": ["Numéro étudiant", "1"],
                "Unnamed: 1": ["Prénom", "Jean"],
                "Unnamed: 2": ["Note", "15"],
            }
        )
        row = find_header_row(df)
        assert row == 0

    def test_promote_header_row(self):
        df = pd.DataFrame(
            {
                "Title": ["Numéro étudiant", "1"],
                "Unnamed: 1": ["Prénom", "Jean"],
                "Unnamed: 2": ["Note", "15"],
            }
        )
        df2 = promote_header_row(df, 0)
        assert "Numéro étudiant" in df2.columns
        assert df2.iloc[0]["Prénom"] == "Jean"


# ============================================================================
# 6. read_file
# ============================================================================


class TestReadFile:
    def test_csv_utf8_comma(self, tmp_path):
        p = tmp_path / "data.csv"
        _write_csv(p, ["A", "B"], [["1", "x"], ["2", "y"]])
        df = read_file(p)
        assert list(df.columns) == ["A", "B"]
        assert len(df) == 2

    def test_csv_semicolon(self, tmp_path):
        p = tmp_path / "data.csv"
        _write_csv(p, ["A", "B"], [["1", "x"]], sep=";")
        df = read_file(p)
        assert list(df.columns) == ["A", "B"]

    def test_csv_latin1(self, tmp_path):
        p = tmp_path / "data.csv"
        _write_csv(p, ["Prénom", "Note"], [["Éloïse", "14,5"]], encoding="latin-1")
        df = read_file(p)
        assert "Prénom" in df.columns
        assert df.iloc[0]["Prénom"] == "Éloïse"

    def test_csv_cp1252_with_apostrophe(self, tmp_path):
        """cp1252 byte 0x92 should decode as U+2019."""
        p = tmp_path / "data.csv"
        with open(p, "wb") as f:
            f.write("Numéro d\u2019identification;Note\n".encode("cp1252"))
            f.write("12345;15\n".encode("cp1252"))
        df = read_file(p)
        col = detect_column(list(df.columns), "id")
        assert col is not None

    def test_xlsx(self, tmp_path):
        p = tmp_path / "data.xlsx"
        pd.DataFrame({"A": ["1"], "B": ["x"]}).to_excel(p, index=False, engine="openpyxl")
        df = read_file(p)
        assert list(df.columns) == ["A", "B"]

    def test_ods(self, tmp_path):
        p = tmp_path / "data.ods"
        pd.DataFrame({"A": ["1"], "B": ["x"]}).to_excel(p, index=False, engine="odf")
        df = read_file(p)
        assert list(df.columns) == ["A", "B"]

    def test_unsupported_extension(self, tmp_path):
        p = tmp_path / "data.json"
        p.write_text("{}")
        with pytest.raises(ValueError, match="Unsupported"):
            read_file(p)

    def test_csv_utf7_encoding(self, tmp_path):
        """UTF-7 encoded CSV should be decoded correctly."""
        p = tmp_path / "data.csv"
        with open(p, "w", encoding="utf-7") as f:
            f.write("Prénom,Nom de famille,Numéro,Note\n")
            f.write("Éloïse,André,12345,15\n")
        df = read_file(p)
        assert "Prénom" in df.columns
        assert df.iloc[0]["Prénom"] == "Éloïse"

    def test_csv_utf7_column_detection(self, tmp_path):
        p = tmp_path / "data.csv"
        with open(p, "w", encoding="utf-7") as f:
            f.write("Prénom,Nom de famille,Numéro,Note\n")
            f.write("Jean,Dupont,12345,15\n")
        df = read_file(p)
        assert detect_column(list(df.columns), "first_name") is not None

    def test_normal_csv_with_plus_not_misdetected(self, tmp_path):
        p = tmp_path / "data.csv"
        _write_csv(p, ["Prénom", "Note"], [["Jean+Pierre", "15"]])
        df = read_file(p)
        assert df.iloc[0]["Prénom"] == "Jean+Pierre"


# ============================================================================
# 7. read_file_sheets — multi-sheet
# ============================================================================


class TestMultiSheet:
    def test_xlsx_single_sheet(self, tmp_path):
        p = tmp_path / "data.xlsx"
        pd.DataFrame({"Numéro étudiant": ["1"], "Note": ["15"]}).to_excel(
            p, index=False, engine="openpyxl"
        )
        sheets = read_file_sheets(p)
        assert len(sheets) == 1

    def test_xlsx_multi_sheet_all_grade_data(self, tmp_path):
        p = tmp_path / "data.xlsx"
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            pd.DataFrame({"Numéro étudiant": ["1"], "Prénom": ["A"], "Note": ["15"]}).to_excel(
                writer, sheet_name="DC-1", index=False
            )
            pd.DataFrame({"Numéro étudiant": ["2"], "Prénom": ["B"], "Note": ["14"]}).to_excel(
                writer, sheet_name="DC-2", index=False
            )
        sheets = read_file_sheets(p)
        assert len(sheets) == 2

    def test_xlsx_skips_non_grade_sheets(self, tmp_path):
        p = tmp_path / "data.xlsx"
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            pd.DataFrame({"Numéro étudiant": ["1"], "Prénom": ["A"], "Note": ["15"]}).to_excel(
                writer, sheet_name="Grades", index=False
            )
            pd.DataFrame({"Foo": ["bar"], "Baz": ["qux"]}).to_excel(
                writer, sheet_name="Metadata", index=False
            )
        sheets = read_file_sheets(p)
        assert len(sheets) == 1
        assert sheets[0][0] == "Grades"

    def test_csv_returns_single_sheet(self, tmp_path):
        p = tmp_path / "data.csv"
        _write_csv(p, ["Numéro étudiant", "Note"], [["1", "15"]])
        sheets = read_file_sheets(p)
        assert len(sheets) == 1
        assert sheets[0][0] == ""


# ============================================================================
# 8. parse_grade
# ============================================================================


class TestParseGrade:
    def test_integer(self):
        assert parse_grade("15").value == 15.0

    def test_float(self):
        assert parse_grade("15.5").value == 15.5

    def test_french_comma(self):
        assert parse_grade("15,5").value == 15.5

    def test_abs(self):
        g = parse_grade("ABS")
        assert g.is_absent

    def test_def(self):
        assert parse_grade("DEF").is_absent

    def test_absence_full_word(self):
        assert parse_grade("Absence").is_absent

    def test_absent_lowercase(self):
        assert parse_grade("abs").is_absent

    def test_empty(self):
        g = parse_grade("")
        assert g.value is None
        assert not g.is_absent

    def test_none(self):
        g = parse_grade(None)
        assert g.value is None

    def test_garbage(self):
        g = parse_grade("foobar")
        assert g.value is None
        assert g.warning is not None


# ============================================================================
# 9. build_master_index
# ============================================================================


class TestBuildMasterIndex:
    def test_basic(self):
        df = _master_df([("S1", "Jean", "Dupont", "j@e.fr")])
        master, _ = build_master_index(df)
        assert "S1" in master.by_id
        assert master.by_id["S1"].first_name == "Jean"

    def test_missing_id_column(self):
        df = pd.DataFrame({"Prénom": ["Jean"], "Nom": ["Dupont"]})
        with pytest.raises(ValueError, match="missing required"):
            build_master_index(df)

    def test_duplicate_id_warning(self):
        df = _master_df(
            [
                ("S1", "Jean", "Dupont", "j@e.fr"),
                ("S1", "Marie", "Curie", "m@e.fr"),
            ]
        )
        master, warnings = build_master_index(df)
        assert any("duplicate" in w.lower() for w in warnings)
        assert master.by_id["S1"].first_name == "Jean"


# ============================================================================
# 10. process_ta_file — basics
# ============================================================================


class TestProcessTaFile:
    def test_basic_id_match(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupont", "15"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert not report.skipped
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0

    def test_french_comma(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupont", "15,5"]],
        )
        process_ta_file(p, master, interactive=False)
        assert master.by_id["12345"].grade == 15.5

    def test_absent_token(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupont", "ABS"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.students_absent == 1
        assert master.by_id["12345"].is_absent

    def test_id_match_with_name_cross_check_mismatch(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupond", "15"]],  # Dupond ≠ Dupont
        )
        report = process_ta_file(p, master, interactive=False)
        # Non-interactive: still proceeds
        assert report.grades_assigned == 1
        assert any("non-interactive" in w.lower() for w in report.warnings)

    def test_name_fallback_when_id_missing(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Prénom", "Nom", "Note"],
            [["Jean", "Dupont", "15"]],
        )
        process_ta_file(p, master, interactive=False)
        assert master.by_id["12345"].grade == 15.0

    def test_unknown_student_skipped(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["99999", "Unknown", "Person", "15"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.grades_assigned == 0

    def test_swapped_first_last_name_recovered(self, tmp_path):
        """TA put surname in Prénom column and vice versa — should still match."""
        master = _build_master()
        p = tmp_path / "ta.csv"
        # No ID column → forces name fallback. Jean Dupont swapped.
        _write_csv(
            p,
            ["Prénom", "Nom", "Note"],
            [["Dupont", "Jean", "15"]],
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0
        assert any("swapping" in w.lower() for w in report.warnings)

    def test_swap_only_used_when_original_fails(self, tmp_path):
        """Original-order match must take precedence over swap."""
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Prénom", "Nom", "Note"],
            [["Jean", "Dupont", "15"]],  # correct order
        )
        report = process_ta_file(p, master, interactive=False)
        assert report.grades_assigned == 1
        # No swap warning — original order matched
        assert not any("swapping" in w.lower() for w in report.warnings)


# ============================================================================
# 11. write_moodle_csv
# ============================================================================


class TestWriteMoodleCsv:
    def test_default_french_columns(self, tmp_path):
        master = _build_master()
        master.by_id["12345"].grade = 15.5
        master.by_id["12346"].is_absent = True

        out = tmp_path / "out.csv"
        write_moodle_csv(master, out)

        df = pd.read_csv(out, dtype=str, keep_default_na=False)
        assert "Numéro d'identification" in df.columns
        assert "Adresse de courriel" in df.columns
        assert "Prénom" in df.columns
        assert "Nom de famille" in df.columns
        assert "Grade" in df.columns

        jean = df[df["Numéro d'identification"] == "12345"].iloc[0]
        assert jean["Grade"] == "15.5"

        marie = df[df["Numéro d'identification"] == "12346"].iloc[0]
        assert marie["Grade"] == "ABS"

    def test_no_grade_is_empty(self, tmp_path):
        master = _build_master()
        out = tmp_path / "out.csv"
        write_moodle_csv(master, out)
        df = pd.read_csv(out, dtype=str, keep_default_na=False)
        assert all(
            df.loc[df["Numéro d'identification"] == sid, "Grade"].values[0] == ""
            for sid in ["12345", "12346", "12347", "12348"]
        )

    def test_custom_exam_name(self, tmp_path):
        master = _build_master()
        master.by_id["12345"].grade = 15.0
        out = tmp_path / "out.csv"
        write_moodle_csv(master, out, exam_name="Partiel Mars 2026")
        df = pd.read_csv(out, dtype=str, keep_default_na=False)
        assert "Partiel Mars 2026" in df.columns
        row = df[df["Numéro d'identification"] == "12345"].iloc[0]
        assert row["Partiel Mars 2026"] == "15"

    def test_custom_id_column(self, tmp_path):
        master = _build_master()
        out = tmp_path / "out.csv"
        write_moodle_csv(master, out, id_column_name="Student Number")
        df = pd.read_csv(out, dtype=str, keep_default_na=False)
        assert "Student Number" in df.columns


# ============================================================================
# 12. Config & resolve_grade_files
# ============================================================================


class TestLoadConfig:
    def test_basic(self, tmp_path):
        p = tmp_path / "c.yaml"
        p.write_text(yaml.dump({"master_file": "m.csv", "grade_files": ["g.csv"]}))
        cfg = load_config(p)
        assert cfg["master_file"] == "m.csv"

    def test_missing_master(self, tmp_path):
        p = tmp_path / "c.yaml"
        p.write_text(yaml.dump({"grade_files": ["g.csv"]}))
        with pytest.raises(ValueError, match="master_file"):
            load_config(p)

    def test_missing_grade_files_and_dir(self, tmp_path):
        p = tmp_path / "c.yaml"
        p.write_text(yaml.dump({"master_file": "m.csv"}))
        with pytest.raises(ValueError):
            load_config(p)


class TestResolveGradeFiles:
    def test_explicit_list(self, tmp_path):
        (tmp_path / "a.csv").write_text("")
        (tmp_path / "b.csv").write_text("")
        result = resolve_grade_files(tmp_path, grade_files=["a.csv", "b.csv"])
        assert len(result) == 2

    def test_directory_scan(self, tmp_path):
        (tmp_path / "a.csv").write_text("")
        (tmp_path / "b.xlsx").write_text("")
        (tmp_path / "ignore.txt.bak").write_text("")
        result = resolve_grade_files(tmp_path, grade_dir=str(tmp_path))
        assert len(result) == 2


# ============================================================================
# 13. Levenshtein and name similarity
# ============================================================================


class TestLevenshteinDistance:
    def test_identical(self):
        assert levenshtein_distance("hello", "hello") == 0

    def test_single_substitution(self):
        assert levenshtein_distance("Dupont", "Dupond") == 1

    def test_single_insertion(self):
        assert levenshtein_distance("Marin", "Martin") == 1

    def test_empty(self):
        assert levenshtein_distance("", "") == 0
        assert levenshtein_distance("abc", "") == 3


class TestNameSimilarityHint:
    def test_identical(self):
        assert "identical" in name_similarity_hint(0, 10)

    def test_likely_typo(self):
        assert "typo" in name_similarity_hint(1, 12)

    def test_very_different(self):
        assert "different" in name_similarity_hint(8, 10)


# ============================================================================
# 14. Interactive prompts
# ============================================================================


class TestPromptColumnChoice:
    def test_select_first(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "1")
        assert prompt_column_choice(["A", "B"], "f.csv") == "A"

    def test_select_skip(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "3")
        assert prompt_column_choice(["A", "B"], "f.csv") is None

    def test_invalid_then_valid(self, monkeypatch):
        inputs = iter(["xyz", "99", "2"])
        monkeypatch.setattr("builtins.input", lambda _: next(inputs))
        assert prompt_column_choice(["A", "B"], "f.csv") == "B"


class TestPromptSheetSelection:
    def test_all(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "all")
        assert prompt_sheet_selection(["X", "Y"], "f.xlsx") == ["X", "Y"]

    def test_none(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "none")
        assert prompt_sheet_selection(["X", "Y"], "f.xlsx") == []

    def test_specific(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "1,3")
        assert prompt_sheet_selection(["X", "Y", "Z"], "f.xlsx") == ["X", "Z"]


class TestPromptNameMismatch:
    def test_yes(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "y")
        assert prompt_name_mismatch("1", "Jean Dupont", "Jean Dupond", "f.csv") is True

    def test_no(self, monkeypatch):
        monkeypatch.setattr("builtins.input", lambda _: "n")
        assert prompt_name_mismatch("1", "Jean Dupont", "Jean Dupond", "f.csv") is False


# ============================================================================
# 15. Ambiguous columns + overrides
# ============================================================================


class TestAmbiguousIdColumn:
    def test_interactive_select(self, tmp_path, monkeypatch):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numero", "Prenom", "Nom", "Numero etudiant", "Note"],
            [["1", "Jean", "Dupont", "12345", "15"]],
        )
        monkeypatch.setattr("builtins.input", lambda _: "2")
        report = process_ta_file(p, master, interactive=True)
        assert master.by_id["12345"].grade == 15.0
        assert report.new_file_overrides is not None
        assert report.new_file_overrides.get("id") == "Numero etudiant"


class TestColumnOverrides:
    def test_id_override_applied(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numero", "Prenom", "Nom", "Numero etudiant", "Note"],
            [["1", "Jean", "Dupont", "12345", "15"]],
        )
        report = process_ta_file(
            p,
            master,
            interactive=False,
            file_overrides={"id": "Numero etudiant"},
        )
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0

    def test_grade_override_applied(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Note /23", "Note /20"],
            [["12345", "18", "15"]],
        )
        report = process_ta_file(
            p,
            master,
            interactive=False,
            file_overrides={"grade": "Note /20"},
        )
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0

    def test_overrides_persisted_to_yaml(self, tmp_path, monkeypatch):
        master_path = tmp_path / "master.csv"
        _write_csv(
            master_path,
            ["Numéro étudiant", "Prénom", "Nom de famille", "Email"],
            [["S001", "Alice", "Martin", "a@e.fr"]],
        )
        ta_path = tmp_path / "ta.csv"
        _write_csv(
            ta_path,
            ["Numéro étudiant", "Note /23", "Note /20"],
            [["S001", "18", "15"]],
        )
        cfg_path = tmp_path / "config.yaml"
        cfg_path.write_text(
            yaml.dump(
                {
                    "master_file": "master.csv",
                    "grade_files": ["ta.csv"],
                    "output_file": "out.csv",
                    "exam_name": "Test",
                    "id_column_name": "ID",
                }
            )
        )
        monkeypatch.setattr("builtins.input", lambda _: "2")
        consolidate(cfg_path, interactive=True)
        with open(cfg_path, encoding="utf-8") as f:
            saved = yaml.safe_load(f)
        assert "column_overrides" in saved
        assert saved["column_overrides"]["ta.csv"]["grade"] == "Note /20"


# ============================================================================
# 16. Name mismatch + name split bug
# ============================================================================


class TestNameMismatchConfirmation:
    def test_interactive_confirm(self, tmp_path, monkeypatch):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupond", "15"]],
        )
        monkeypatch.setattr("builtins.input", lambda _: "y")
        report = process_ta_file(p, master, interactive=True)
        assert report.grades_assigned == 1
        assert "12345" in report.new_name_confirmations

    def test_interactive_reject(self, tmp_path, monkeypatch):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Pierre", "Martin", "15"]],
        )
        monkeypatch.setattr("builtins.input", lambda _: "n")
        report = process_ta_file(p, master, interactive=True)
        assert report.grades_assigned == 0

    def test_saved_confirmation_skips_prompt(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Jean", "Dupond", "15"]],
        )
        report = process_ta_file(
            p,
            master,
            interactive=True,
            name_confirmations=["12345"],
        )
        assert report.grades_assigned == 1
        assert any("previously confirmed" in w.lower() for w in report.warnings)


class TestNameSplitMismatch:
    def test_different_split_same_full_name_no_prompt(self, tmp_path):
        """First/last split differs but full name identical → no prompt."""
        df = _master_df([("12345", "Mohamed Ayoub", "Mebarki", "ma@e.fr")])
        master, _ = build_master_index(df)
        p = tmp_path / "ta.csv"
        _write_csv(
            p,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [["12345", "Mohamed", "Ayoub Mebarki", "15"]],
        )
        # No monkeypatch — must not prompt
        report = process_ta_file(p, master, interactive=True)
        assert report.grades_assigned == 1
        assert master.by_id["12345"].grade == 15.0


# ============================================================================
# 17. Sheet selection
# ============================================================================


class TestSheetSelectionIntegration:
    def test_saved_selection_reused(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Current"
        ws1.append(["Numéro étudiant", "Prénom", "Note"])
        ws1.append(["12345", "Jean", "15"])
        ws2 = wb.create_sheet("Future")
        ws2.append(["Numéro étudiant", "Prénom", "Note"])
        ws2.append(["12345", "Jean", ""])
        wb.save(p)

        report = process_ta_file(
            p,
            master,
            interactive=True,
            file_overrides={"selected_sheets": ["Current"]},
        )
        assert report.grades_assigned == 1
        assert any("saved sheet selection" in w.lower() for w in report.warnings)

    def test_skip_none_sheets(self, tmp_path, monkeypatch):
        master = _build_master()
        p = tmp_path / "ta.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.append(["Numéro étudiant", "Prénom", "Note"])
        ws1.append(["12345", "Jean", "15"])
        ws2 = wb.create_sheet("S2")
        ws2.append(["Numéro étudiant", "Prénom", "Note"])
        ws2.append(["12346", "Marie", "14"])
        wb.save(p)

        monkeypatch.setattr("builtins.input", lambda _: "none")
        report = process_ta_file(p, master, interactive=True)
        assert report.skipped


class TestMultiSheetOverridePropagation:
    def test_grade_choice_carries_across_sheets(self, tmp_path, monkeypatch):
        master = _build_master()
        p = tmp_path / "ta.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "DC-1"
        ws1.append(["Numéro étudiant", "Prénom", "Note /23", "Note /20"])
        ws1.append(["12345", "Jean", "18", "15"])
        ws2 = wb.create_sheet("DC-2")
        ws2.append(["Numéro étudiant", "Prénom", "Note /23", "Note /20"])
        ws2.append(["12346", "Marie", "17", "14"])
        wb.save(p)

        # Sheet selection: "all", then grade column for first sheet: "2"
        responses = iter(["all", "2"])
        monkeypatch.setattr("builtins.input", lambda _: next(responses))
        report = process_ta_file(p, master, interactive=True)

        assert report.grades_assigned == 2
        assert master.by_id["12345"].grade == 15.0
        assert master.by_id["12346"].grade == 14.0
        assert report.new_file_overrides is not None
        assert "selected_sheets" in report.new_file_overrides
        assert "sheet_columns" in report.new_file_overrides

    def test_saved_per_sheet_overrides(self, tmp_path):
        master = _build_master()
        p = tmp_path / "ta.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "DC-1"
        ws1.append(["Numéro étudiant", "Prénom", "Note /23", "Note /20"])
        ws1.append(["12345", "Jean", "18", "15"])
        ws2 = wb.create_sheet("DC-2")
        ws2.append(["Numéro étudiant", "Prénom", "Note /23", "Note /20"])
        ws2.append(["12346", "Marie", "17", "14"])
        wb.save(p)

        report = process_ta_file(
            p,
            master,
            interactive=True,
            file_overrides={
                "selected_sheets": ["DC-1", "DC-2"],
                "sheet_columns": {
                    "DC-1": {"grade": "Note /20"},
                    "DC-2": {"grade": "Note /20"},
                },
            },
        )
        assert report.grades_assigned == 2


# ============================================================================
# 18. Integration
# ============================================================================


class TestIntegration:
    def test_end_to_end(self, tmp_path):
        master_path = tmp_path / "master.csv"
        _write_csv(
            master_path,
            ["Numéro étudiant", "Prénom", "Nom de famille", "Email"],
            [
                ["S001", "Alice", "Martin", "a@e.fr"],
                ["S002", "Bob", "Bernard", "b@e.fr"],
                ["S003", "Carol", "Petit", "c@e.fr"],
            ],
        )

        ta1 = tmp_path / "g1.csv"
        _write_csv(
            ta1,
            ["Numéro étudiant", "Prénom", "Nom", "Note"],
            [
                ["S001", "Alice", "Martin", "16"],
                ["S002", "Bob", "Bernard", "ABS"],
            ],
        )

        ta2 = tmp_path / "g2.xlsx"
        pd.DataFrame(
            {
                "Numéro étudiant": ["S003"],
                "Prénom": ["Carol"],
                "Nom": ["Petit"],
                "Note": ["13.5"],
            }
        ).to_excel(ta2, index=False, engine="openpyxl")

        cfg_path = tmp_path / "config.yaml"
        cfg_path.write_text(
            yaml.dump(
                {
                    "master_file": "master.csv",
                    "grade_files": ["g1.csv", "g2.xlsx"],
                    "output_file": "out.csv",
                    "exam_name": "Partiel",
                    "id_column_name": "Numéro d'identification",
                }
            )
        )

        master, reports = consolidate(cfg_path, interactive=False)
        assert all(not r.skipped for r in reports)
        assert master.by_id["S001"].grade == 16.0
        assert master.by_id["S002"].is_absent
        assert master.by_id["S003"].grade == 13.5

        out = tmp_path / "out.csv"
        df = pd.read_csv(out, dtype=str, keep_default_na=False)
        assert "Partiel" in df.columns
        assert "Numéro d'identification" in df.columns
        assert df.loc[df["Numéro d'identification"] == "S002", "Partiel"].values[0] == "ABS"